import sys
import os
import openpyxl
from cheap_repr import cheap_repr
from openpyxl.styles import Font
from contextlib import contextmanager

def deleteLastLineJson():
    '''
    helping function. function delete one line in json file
    '''
    fd = open("tracer_output.json", "r")
    d = fd.read()
    fd.close()
    m = d.split("\n")
    s = "\n".join(m[:-1])
    fd = open("tracer_output.json", "w+")
    for i in range(len(s)):
        fd.write(s[i])
    fd.close()

class DefaultOutput:
    '''
    Default output format class.
    '''
    def __init__(self, counter):
        '''
        init function.
        initialization of variables
        :param counter: number of tabs printed. at start = 0.
        '''
        self.counter = counter

    def entry(self, separator, func_name, frame):
        '''
        print "call" function informations
        :param separator: tab
        :param func_name: name of function
        :param frame: is the current stack frame.
        '''
        print(f'{separator * self.counter}-->{func_name}({cheap_repr(frame.f_locals)}) line: {frame.f_lineno}')
        self.counter += 1

    def exit(self, separator, func_name, frame, arg):
        '''
        print "return" function informations
        :param separator: tab
        :param func_name: name of function
        :param frame: is the current stack frame
        :param arg: return value
        '''
        self.counter -= 1
        print(separator * self.counter, "<-- ", func_name, "(", cheap_repr(frame.f_locals), ") res =", cheap_repr(arg), "line: ",frame.f_lineno)

    def exception(self, frame, func_name):
        '''
        print "exception" informations
        :param func_name: name of function
        :param frame: is the current stack frame
        '''
        print("Tracing exception on line ", frame.f_lineno, "of ", func_name, "(", cheap_repr(frame.f_locals), ")")

class ColorOutput:
    '''
    Color output format class.
    '''
    def __init__(self, counter):
        '''
        init function.
        initialization of variables
        :param counter: number of tabs printed. at start = 0.
        '''
        self.counter = counter

    def entry(self, separator, func_name, frame):
        '''
        print "call" function informations
        :param separator: tab
        :param func_name: name of function
        :param frame: is the current stack frame.
        '''
        print(f'{bcolors.OK}{separator * self.counter}-->{func_name}({cheap_repr(frame.f_locals)}) line: {frame.f_lineno} {bcolors.RESET}')
        self.counter += 1

    def exit(self, separator, func_name, frame, arg):
        '''
        print "return" function informations
        :param separator: tab
        :param func_name: name of function
        :param frame: is the current stack frame
        :param arg: return value
        '''
        self.counter -= 1
        print(separator * self.counter, bcolors.WARNING + "<-- ", func_name, "(", cheap_repr(frame.f_locals), ") res =", cheap_repr(arg),"line: ", frame.f_lineno, bcolors.RESET)

    def exception(self, frame, func_name):
        '''
        print "exception" informations
        :param func_name: name of function
        :param frame: is the current stack frame
        '''
        print(bcolors.FAIL + "Tracing exception on line ", frame.f_lineno, "of ", func_name, "(", cheap_repr(frame.f_locals), ")", bcolors.RESET)

class FileOutput:
    '''
    output in to .txt file format class.
    '''
    def __init__(self, counter):
        '''
        init function.
        initialization of variables
        open .txt file or create it if it doesn't exist
        :param counter: number of tabs printed. at start = 0.
        '''
        self.counter = counter
        self.f = open("Pytrace_output.txt", "w")

    def entry(self, separator, func_name, frame):
        '''
        print "call" function informations
        :param separator: tab
        :param func_name: name of function
        :param frame: is the current stack frame.
        '''
        if self.counter==0:
            self.f = open("Pytrace_output.txt", "a")
            self.f.write(f'{separator * self.counter}-->{func_name}({cheap_repr(frame.f_locals)}) line: {frame.f_lineno}\n')
        else:
            self.f = open("Pytrace_output.txt", "a")
            self.f.write(f'{separator * self.counter}-->{func_name}({cheap_repr(frame.f_locals)}) line: {frame.f_lineno}\n')
        self.counter += 1


    def exit(self, separator, func_name, frame, arg):
        '''
        print "return" function informations
        :param separator: tab
        :param func_name: name of function
        :param frame: is the current stack frame
        :param arg: return value
        '''
        self.counter -= 1
        self.f = open("Pytrace_output.txt", "a")
        self.f.write(f'{separator * self.counter}<--{func_name}({cheap_repr(frame.f_locals)}) res = {cheap_repr(arg)} line: {frame.f_lineno}\n')
        if self.counter==0:
            self.f.close()

    def exception(self, frame, func_name):
        self.f.write(f'Tracing exception on line {frame.f_lineno} of {func_name}({cheap_repr(frame.f_locals)})\n')

class JsonOutput:
    '''
    Json output format class.
    '''
    def __init__(self, counter):
        '''
        init function.
        initialization of variables
        open .json file or create it if it doesn't exist
        :param counter: number of tabs printed. at start = 0.
        '''
        self.counter = counter
        self.f = open("tracer_output.json", "w")
        self.counteroffunc = 1

    def entry(self, separator, func_name, frame):
        '''
        print "call" function informations
        :param separator: tab
        :param func_name: name of function
        :param frame: is the current stack frame.
        '''
        if(func_name != "__exit__" and func_name != "start_tracing"):

            if self.counter==0:
                #delete last line 3*
                if self.counteroffunc != 1:
                    deleteLastLineJson()
                    deleteLastLineJson()
                    deleteLastLineJson()


                self.f = open("tracer_output.json", "a")
                if self.counteroffunc == 1:
                    self.f.write(f'{"{"}\n')
                else:
                    self.f.write(f',\n')
                if self.counteroffunc == 1:
                    self.f.write(f'\"output\": {"["}\n')
                self.f.write(f'{separator}{"{"}\n')


                self.f.write(f'{separator * 2}\"func\": "func{"("}{self.counteroffunc}{")"}\",\n')


                self.f.write(f'{separator * 2}\"name\": \"{sys.argv[0]}\",\n')
                self.f.write(f'{separator * 2}\"tracing\": {"["}\n')
                self.f.write(f'{separator * 3}{"{"}\"type\": \"-->\", \"func_name\": \"{func_name}\", \"variables\": {"["}\"{cheap_repr(frame.f_locals)}\"{"]"}, \"line\": {frame.f_lineno}{"}"}')


            else:
                self.f = open("tracer_output.json", "a")
                self.f.write(f',\n{separator * 3}{"{"}\"type\": \"-->\", \"func_name\": \"{func_name}\", \"variables\": {"["}\"{cheap_repr(frame.f_locals)}\"{"]"}, \"line\": {frame.f_lineno}{"}"}')

                #self.f.write(f'{separator * self.counter}-->{func_name}({str(frame.f_locals)}) line: {frame.f_lineno}\n')
            self.counter += 1


    def exit(self, separator, func_name, frame, arg):
        '''
        print "return" function informations
        :param separator: tab
        :param func_name: name of function
        :param frame: is the current stack frame
        :param arg: return value
        '''
        self.counter -= 1
        self.f = open("tracer_output.json", "a")
        if arg == None:
            self.f.write(
                f',\n{separator * 3}{"{"}\"type\": \"<--\", \"func_name\": \"{func_name}\", \"variables\": {"["}\"{cheap_repr(frame.f_locals)}\"{"]"},\"result\": \"None\", \"line\": {frame.f_lineno}{"}"}')
        else:
            self.f.write(f',\n{separator * 3}{"{"}\"type\": \"<--\", \"func_name\": \"{func_name}\", \"variables\": {"["}\"{cheap_repr(frame.f_locals)}\"{"]"},\"result\": {cheap_repr(arg)}, \"line\": {frame.f_lineno}{"}"}')

        if self.counter==0:
            self.counteroffunc += 1
            self.f.write(f'\n{separator * 3}{"]"}')
            self.f.write(f'\n{separator * 2}{"}"}')
            self.f.write(f'\n{separator}{"]"}\n')
            self.f.write(f'{"}"}\n')
            self.f.close()

    def exception(self, frame, func_name):
        '''
        print "exception" informations
        :param func_name: name of function
        :param frame: is the current stack frame
        '''
        self.f.write(f',\n\t\t\t{"{"}\"type\": \"exception\", \"func_name\": \"{func_name}\", \"variables\": {"["}\"{cheap_repr(frame.f_locals)}\"{"]"}, \"line\": {frame.f_lineno}{"}"}')

class ExcelOutput:
    '''
    output in to excel file format class.
    '''
    def __init__(self, counter):
        '''
        init function.
        initialization of variables
        ctreate and open .xlsx file or remove it if it exist.
        setup excel file.
        :param counter: number of tabs printed. at start = 0.
        '''
        if os.path.exists("Pytrace.xlsx"):
            os.remove("Pytrace.xlsx")

        self.counter = counter
        self.row = 1
        self.column = 1

        wb = openpyxl.Workbook()
        sheet = wb.active

        sheet.column_dimensions['A'].width = 17

        sheet.cell(row=self.row, column=self.column).value = 'Name of function'
        sheet.cell(row=self.row, column=self.column).font = Font(bold=True)
        self.column += 1
        sheet.cell(row=self.row, column=self.column).value = 'Event type'
        sheet.cell(row=self.row, column=self.column).font = Font(bold=True)
        self.column += 1
        sheet.cell(row=self.row, column=self.column).value = 'Variables'
        sheet.cell(row=self.row, column=self.column).font = Font(bold=True)
        self.column += 1
        sheet.cell(row=self.row, column=self.column).value = 'Line'
        sheet.cell(row=self.row, column=self.column).font = Font(bold=True)
        self.column += 1
        sheet.cell(row=self.row, column=self.column).value = 'Result'
        sheet.cell(row=self.row, column=self.column).font = Font(bold=True)

        wb.save("Pytrace.xlsx")
        self.counteroffunc = 1

    def entry(self, separator, func_name, frame):
        '''
        print "call" function informations
        :param separator: tab
        :param func_name: name of function
        :param frame: is the current stack frame.
        '''
        if(func_name != "__exit__" and func_name != "start_tracing"):
            self.row = self.row + 1
            self.column = 1
            wb = openpyxl.load_workbook("Pytrace.xlsx")
            sheet = wb.active

            sheet.cell(row=self.row, column=self.column).value = str(func_name)
            self.column += 1
            sheet.cell(row=self.row, column=self.column).value = 'Entrance'
            self.column += 1
            sheet.cell(row=self.row, column=self.column).value = cheap_repr(frame.f_locals)
            self.column += 1
            sheet.cell(row=self.row, column=self.column).value = str(frame.f_lineno)

            wb.save("Pytrace.xlsx")

            self.counter += 1



    def exit(self, separator, func_name, frame, arg):
        '''
        print "return" function informations
        :param separator: tab
        :param func_name: name of function
        :param frame: is the current stack frame
        :param arg: return value
        '''
        self.row = self.row + 1
        self.column = 1
        wb = openpyxl.load_workbook("Pytrace.xlsx")
        sheet = wb.active

        sheet.cell(row=self.row, column=self.column).value = str(func_name)
        self.column += 1
        sheet.cell(row=self.row, column=self.column).value = 'Exit'
        self.column += 1
        sheet.cell(row=self.row, column=self.column).value = cheap_repr(frame.f_locals)
        self.column += 1
        sheet.cell(row=self.row, column=self.column).value = str(frame.f_lineno)
        self.column += 1

        if arg == None:
            sheet.cell(row=self.row, column=self.column).value = 'None'
        else:
            sheet.cell(row=self.row, column=self.column).value = cheap_repr(arg)
        wb.save("Pytrace.xlsx")
        if self.counter == 0:
            wb.save("Pytrace.xlsx")
        self.counter -= 1

    def exception(self, frame, func_name):
        '''
        print "exception" informations
        :param func_name: name of function
        :param frame: is the current stack frame
        '''
        self.row = self.row + 1
        self.column = 1
        wb = openpyxl.load_workbook("Pytrace.xlsx")
        sheet = wb.active

        sheet.cell(row=self.row, column=self.column).value = str(func_name)
        self.column += 1
        sheet.cell(row=self.row, column=self.column).value = 'Excetion'
        self.column += 1
        sheet.cell(row=self.row, column=self.column).value = str(frame.f_locals)
        self.column += 1
        sheet.cell(row=self.row, column=self.column).value = str(frame.f_lineno)
        wb.save("Pytrace.xlsx")

class bcolors:
    '''
    set colors to variables
    '''
    OK = '\033[92m' #GREEN
    WARNING = '\033[93m' #YELLOW
    FAIL = '\033[91m' #RED
    RESET = '\033[0m' #RESET COLO


def trace_calls_and_returns(frame, event, arg):
    '''
    tracing function
    :param frame: frame is the current stack frame
    :param event: (String)typ of tracing event (call, return or exception)
    :param arg: depends on the event type. return value of return event or None.
    '''
    co = frame.f_code
    func_name = co.co_name
    separator = '\t'

    if event == 'call':
        OUTPUT.entry(separator, func_name, frame)
        return trace_calls_and_returns

    elif event == 'return':
        OUTPUT.exit(separator, func_name, frame, arg)

    elif event == 'exception':
        OUTPUT.exception(frame, func_name)

    return



@contextmanager
def start_tracing():
    '''
    function print info and start tracing calling sys.settrace() with argument
    The argument is tracing function.
    '''
    get_trace = sys.gettrace()
    print("This is the name of the script: ", sys.argv[0])
    print("Number of arguments: ", len(sys.argv))
    print("The arguments are: ", str(sys.argv))
    sys.settrace(trace_calls_and_returns)
    yield
    sys.settrace(get_trace)

OUTPUT = DefaultOutput(0)
# OUTPUT = ColorOutput(0)
# OUTPUT = FileOutput(0)
# OUTPUT = ExcelOutput(0)
# OUTPUT = JsonOutput(0)
